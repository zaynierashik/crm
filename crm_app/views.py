import io
import json
import random
import pandas as pd

from django.shortcuts import render, redirect, get_object_or_404
from django.http import *
from django.utils.timezone import now
from datetime import datetime, timezone
from django.urls import reverse
from django.db.models import Sum, Count
from .models import *
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.contrib.auth.hashers import check_password, make_password
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth import login as auth_login, authenticate, logout as auth_logout, update_session_auth_hash
from openpyxl import Workbook # type: ignore
from openpyxl.styles import Alignment, Font, Border, Side # type: ignore
from django.utils.dateformat import DateFormat

from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags

from django.http import JsonResponse

# Login Page
def login(request):
    if 'staff_id' in request.session:
        return redirect('dashboard')
    
    if request.method == 'POST':
        return login_authentication(request)
    
    return render(request, 'login.html')

# Login Authentication
def login_authentication(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        try:
            user = Staff.objects.get(email=email)
        except Staff.DoesNotExist:
            error_message = "Admin/Staff does not exist!"
            return HttpResponseRedirect(reverse('login') + f'?error_message={error_message}')

        if not check_password(password, user.password):
            error_message = "Invalid password!"
            return HttpResponseRedirect(reverse('login') + f'?error_message={error_message}')

        if not user.status:
            error_message = "Your account is inactive. Please contact support."
            return HttpResponseRedirect(reverse('login') + f'?error_message={error_message}')

        request.session['staff_id'] = user.id
        request.session['staff_email'] = user.email
        request.session['staff_full_name'] = user.full_name

        request.session.set_expiry(7200)
        return redirect('dashboard')

    return render(request, 'login.html')

# Logout
def logout(request):
    request.session.flush()
    return redirect('login')

# Dashboard
def dashboard(request):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = Staff.objects.get(id=staff_id)
    
    companies = Company.objects.order_by('company_name')
    contacts = Contact.objects.select_related('company').all()
    sectors = Sector.objects.values('id', 'sector_name').order_by('sector_name')
    services = Service.objects.values('id', 'service_name').distinct().order_by('service_name')
    brands = Brand.objects.values('id', 'brand_name').distinct().order_by('brand_name')
    products = Product.objects.values('id', 'product_name').distinct().order_by('product_name')
    cities = Company.objects.values_list('city', flat=True).distinct().order_by('city')
    staffs = Staff.objects.all()

    company_count = Company.objects.count()
    pending_contracts = Requirement.objects.filter(progress='Pipeline').count()

    total_revenue = 0

    progress_counts = Requirement.objects.values('progress').annotate(count=Count('id')).filter(progress__in=['Initiated', 'Pipeline', 'Completed'])
    progress_labels = [entry['progress'] for entry in progress_counts]
    progress_data = [entry['count'] for entry in progress_counts]

    company_trends = (Company.objects.values('date').annotate(count=Count('id')).order_by('date'))
    dates = [DateFormat(entry['date']).format('Y-m-d') for entry in company_trends]
    counts = [entry['count'] for entry in company_trends]

    context = {'companies': companies, 'contacts': contacts, 'sectors': sectors, 'services': services, 'brands': brands, 'products': products, 
               'cities': cities, 'staffs': staffs, 'pending_contracts': pending_contracts, 'total_revenue': round(total_revenue, 2), 
               'company_count': company_count, 'user': user, 'progress_labels': progress_labels, 'progress_data': progress_data, 'chart_dates': dates, 'chart_counts': counts}

    return render(request, 'index.html', context)

# Company Page
def company(request):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = Staff.objects.get(id=staff_id)

    COUNTRY_CHOICES = [
        ('Nepal', 'Nepal'),
        ('India', 'India'),
        ('Singapore', 'Singapore'),
    ]

    ROLES = [
        ('Client', 'Client'),
        ('Partner', 'Partner')
    ]

    companies = Company.objects.order_by('company_name')
    contacts = Contact.objects.select_related('company').all()
    sectors = Sector.objects.values('id', 'sector_name').order_by('sector_name')
    services = Service.objects.values('id', 'service_name').distinct().order_by('service_name')
    brands = Brand.objects.values('id', 'brand_name').distinct().order_by('brand_name')
    products = Product.objects.values('id', 'product_name').distinct().order_by('product_name')
    cities = Company.objects.values_list('city', flat=True).distinct().order_by('city')
    staffs = Staff.objects.all()

    partner_companies = Company.objects.filter(role='Partner').order_by('company_name')
    active_companies = Company.objects.filter(status=True).order_by('company_name')
    inactive_companies = Company.objects.filter(status=False).order_by('company_name')

    company_count = Company.objects.count()
    initiated_count = Requirement.objects.filter(status="Initiated").count()
    pipeline_count = Requirement.objects.filter(status="Pipeline").count()
    completed_count = Requirement.objects.filter(status="Completed").count()

    city_filter = request.GET.get('city', None)
    status_filter = request.GET.get('status', 'active')

    if city_filter:
        active_companies = active_companies.filter(city=city_filter)
        inactive_companies = inactive_companies.filter(city=city_filter)

    if status_filter:
        status_value = status_filter.lower() == 'active'
        if status_value:
            active_companies = active_companies.filter(status=True)
        else:
            inactive_companies = inactive_companies.filter(status=False)

    paginator_active = Paginator(active_companies, 10)
    page_active = request.GET.get('page_active')

    try:
        active_companies_page = paginator_active.page(page_active)
    except PageNotAnInteger:
        active_companies_page = paginator_active.page(1)
    except EmptyPage:
        active_companies_page = paginator_active.page(paginator_active.num_pages)

    paginator_inactive = Paginator(inactive_companies, 10)
    page_inactive = request.GET.get('page_inactive')

    try:
        inactive_companies_page = paginator_inactive.page(page_inactive)
    except PageNotAnInteger:
        inactive_companies_page = paginator_inactive.page(1)
    except EmptyPage:
        inactive_companies_page = paginator_inactive.page(paginator_inactive.num_pages)

    context = {'country_choices': COUNTRY_CHOICES, 'roles': ROLES, 'companies': companies, 'contacts': contacts, 'sectors': sectors, 'services': services, 'brands': brands, 'products': products, 'cities': cities, 'staffs': staffs, 'company_count': company_count,
                'initiated_count': initiated_count, 'pipeline_count': pipeline_count, 'completed_count': completed_count, 'partner_companies': partner_companies, 'active_companies': active_companies_page, 'inactive_companies': inactive_companies_page,
                'paginator_active': paginator_active, 'paginator_inactive': paginator_inactive, 'page_active': page_active, 'page_inactive': page_inactive, 'user': user}

    return render(request, 'company.html', context)

# Add New Company
def add_newcompany(request):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    staff = Staff.objects.get(id=staff_id)

    if request.method == 'POST':
        company_name = request.POST.get('company-name')
        role = request.POST.get('role') or 'Client'

        if Company.objects.filter(company_name=company_name).exists():
            error_message = "Company with this name already exists!"
            return HttpResponseRedirect(reverse('company') + f'?add-company&error_message={error_message}&company_name={company_name}')

        sector_id = request.POST.get('sector')
        if sector_id:
            try:
                sector_id = int(sector_id)
                sector = Sector.objects.get(id=sector_id)
            except (ValueError, Sector.DoesNotExist):
                sector = None
        else:
            sector = None
        
        address = request.POST.get('address')
        city = request.POST.get('city')
        state = request.POST.get('state')
        country = request.POST.get('country')
        via = request.POST.get('via')
        website = request.POST.get('website')
        connection = request.POST.get('connection')

        if via == 'Referral':
            referral = request.POST.get('referral')
        else:
            referral = None

        # Handle partner company assignment when via is Partner
        partner_company = None
        if via == 'Partner':
            partner_id = request.POST.get('partner')  # Get partner company ID from the form
            if partner_id:
                try:
                    partner_company = Company.objects.get(id=partner_id)
                except Company.DoesNotExist:
                    partner_company = None
            else:
                partner_company = None

        company = Company(company_name=company_name, role=role, sector=sector, address=address, city=city, state=state, country=country, via=via, referral_name=referral, partner_company=partner_company, website=website, connection=connection, created_by=staff)
        company.save()

        contact_names = request.POST.getlist('contact-name[]')
        designations = request.POST.getlist('designation[]')
        emails = request.POST.getlist('email-address[]')
        contact_numbers = request.POST.getlist('contact-number[]')
        dobs = request.POST.getlist('date[]')
        religions = request.POST.getlist('religion[]')

        for i in range(len(contact_names)):
            if contact_names[i]:
                dob = dobs[i] if dobs[i] else None
                contact_person = Contact(company=company, contact_name=contact_names[i], designation=designations[i], email=emails[i], phone_number=contact_numbers[i], DOB=dob, religion=religions[i])
                contact_person.save()
                
        return redirect(reverse('company'))
    else:
        return HttpResponse("Form Submission Error!")

# Contract Page
def contract(request):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = Staff.objects.get(id=staff_id)
    
    requirements = Requirement.objects.select_related('company').all().order_by('-date')
    services = Service.objects.values('id', 'service_name').order_by('service_name')
    brands = Brand.objects.values('id', 'brand_name').order_by('brand_name')
    products = Product.objects.values('id', 'product_name').order_by('product_name')

    total_requirement = Requirement.objects.count()
    pipeline_count = Requirement.objects.filter(progress="Pipeline").count()
    completed_count = Requirement.objects.filter(progress="Completed").count()

    paginator = Paginator(requirements, 10)
    page = request.GET.get('requirements_page')

    try:
        requirements_page = paginator.page(page)
    except PageNotAnInteger:
        requirements_page = paginator.page(1)
    except EmptyPage:
        requirements_page = paginator.page(paginator.num_pages)

    context = {'user': user, 'requirements': requirements, 'services': services, 'brands': brands, 'products': products, 'total_requirement': total_requirement, 
               'pipeline_count': pipeline_count, 'completed_count': completed_count, 'requirement_count': requirements.count(), 'requirements_page': requirements_page, 'paginator': paginator}
    
    return render(request, 'contract.html', context)

# Tasks Page
def task(request):
    if 'staff_id' not in request.session:
        return redirect('login')

    staff_id = request.session.get('staff_id')
    user = Staff.objects.get(id=staff_id)

    task_status = request.GET.get('status', None)

    if user.role == 'Staff':
        tasks = Task.objects.filter(assigned_to=user).order_by('-created_at')
    else:
        tasks = Task.objects.all().order_by('-created_at')

    if task_status:
        tasks = tasks.filter(status=task_status)

    total_tasks = tasks.count()

    staffs = Staff.objects.filter(role='Staff')

    # Separate tasks into pending and completed
    pending_tasks = tasks.filter(status__in=['Pending', 'In Progress']).order_by('-created_at')
    completed_tasks = tasks.filter(status='Completed').order_by('-created_at')

    pending_tasks_count = pending_tasks.count()
    completed_tasks_count = completed_tasks.count()

    # Total tasks pagination
    total_paginator = Paginator(tasks, 10)
    page = request.GET.get('page')
    try:
        total_tasks_page = total_paginator.page(page)
    except PageNotAnInteger:
        total_tasks_page = total_paginator.page(1)
    except EmptyPage:
        total_tasks_page = total_paginator.page(total_paginator.num_pages)

    # Pending tasks pagination
    pending_paginator = Paginator(pending_tasks, 10)
    pending_page = request.GET.get('pending_page')
    try:
        pending_tasks_page = pending_paginator.page(pending_page)
    except PageNotAnInteger:
        pending_tasks_page = pending_paginator.page(1)
    except EmptyPage:
        pending_tasks_page = pending_paginator.page(pending_paginator.num_pages)

    # Completed tasks pagination
    completed_paginator = Paginator(completed_tasks, 10)
    completed_page = request.GET.get('completed_page')
    try:
        completed_tasks_page = completed_paginator.page(completed_page)
    except PageNotAnInteger:
        completed_tasks_page = completed_paginator.page(1)
    except EmptyPage:
        completed_tasks_page = completed_paginator.page(completed_paginator.num_pages)

    context = {
        'user': user, 'staffs': staffs, 'tasks': total_tasks_page, 'pending_tasks': pending_tasks_page, 'completed_tasks': completed_tasks_page, 'pending_tasks_count': pending_tasks_count,
        'completed_tasks_count': completed_tasks_count, 'total_tasks': total_tasks_page, 'pending_paginator': pending_paginator, 'completed_paginator': completed_paginator}

    return render(request, 'task.html', context)

# Assign Task
def assign_newtask(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        assigned_to_id = request.POST.get('full-name')
        priority = request.POST.get('priority')
        due_date = request.POST.get('date')
        
        assigned_to = Staff.objects.get(id=assigned_to_id)
        
        task = Task.objects.create(title=title, description=description, due_date=due_date, priority=priority, assigned_to=assigned_to)
        return redirect('task')

    return render(request, 'task.html')

# Requirement Page
def requirement(request):
    # if 'user_id' not in request.session:
    #     return redirect('user')
    
    # user_id = request.session.get('user_id')
    # user = get_object_or_404(User, id=user_id)

    # company = Company.objects.filter(created_by=user).first()

    # services = Service.objects.values('id', 'service_name').distinct().order_by('service_name')
    # brands = Brand.objects.values('id', 'brand_name').distinct().order_by('brand_name')
    # products = Product.objects.values('id', 'product_name').distinct().order_by('product_name')
    # requests = Request.objects.filter(company=company).order_by('-date')
    # requirements = Requirement.objects.filter(company=company).order_by('-date')

    # total_request = Request.objects.filter(company=company).count()
    # total_requirement = Requirement.objects.filter(company=company).count()
    # pipeline_count = Requirement.objects.filter(company=company, progress="Pipeline").count()
    # completed_count = Requirement.objects.filter(company=company, progress="Completed").count()

    # paginator = Paginator(requirements, 10)
    # page = request.GET.get('requirements_page')

    # try:
    #     requirements_page = paginator.page(page)
    # except PageNotAnInteger:
    #     requirements_page = paginator.page(1)
    # except EmptyPage:
    #     requirements_page = paginator.page(paginator.num_pages)

    # context = {'services': services, 'brands': brands, 'products': products, 'requests': requests, 'requirements': requirements, 'user': user, 'total_request': total_request, 'total_requirement': total_requirement, 
    #            'pipeline_count': pipeline_count, 'completed_count': completed_count, 'requirement_count': requirements.count(), 'requirements_page': requirements_page, 'paginator': paginator}

    return render(request, 'requirement.html')

# Staff Page
def staff(request):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_role = Staff.objects.filter(id=request.session['staff_id']).first()
    
    # Check if the current staff member exists and is an admin
    if not staff_role or staff_role.role.lower() != 'admin':
        return redirect('login')

    staffs = Staff.objects.order_by('full_name')
    employee_count = Staff.objects.count()
    admin_count = Staff.objects.filter(role="Admin").count()
    staff_count = Staff.objects.filter(role="Staff").count()
    inactive_count = Staff.objects.filter(status=False).count()
    paginator = Paginator(staffs, 10)
    page = request.GET.get('page')

    try:
        staffs_page = paginator.page(page)
    except PageNotAnInteger:
        staffs_page = paginator.page(1)
    except EmptyPage:
        staffs_page = paginator.page(paginator.num_pages)

    context = {'staffs': staffs, 'employee_count': employee_count, 'admin_count': admin_count, 'staff_count': staff_count, 'inactive_count': inactive_count, 'staffs': staffs_page, 'paginator': paginator, 'page_obj': staffs_page}

    return render(request, 'staff.html', context)

# Transaction Page
def transaction(request):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = Staff.objects.get(id=staff_id)

    companies = Company.objects.order_by('company_name')
    contacts = Contact.objects.select_related('company').all()
    sectors = Sector.objects.values('id', 'sector_name').order_by('sector_name')
    services = Service.objects.values('id', 'service_name').distinct().order_by('service_name')
    brands = Brand.objects.values('id', 'brand_name').distinct().order_by('brand_name')
    products = Product.objects.values('id', 'product_name').distinct().order_by('product_name')
    cities = Company.objects.values_list('city', flat=True).distinct().order_by('city')
    staffs = Staff.objects.all()

    context = {'companies': companies, 'contacts': contacts, 'sectors': sectors, 'services': services, 'brands': brands, 'products': products, 
                'cities': cities, 'staffs': staffs, 'user': user}

    return render(request, 'minute.html', context)

# Sector Page
def sector(request):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = Staff.objects.get(id=staff_id)

    sectors = Sector.objects.order_by('sector_name')
    sector_count = Sector.objects.count()
    paginator = Paginator(sectors, 10)
    page = request.GET.get('page')

    try:
        sectors_page = paginator.page(page)
    except PageNotAnInteger:
        sectors_page = paginator.page(1)
    except EmptyPage:
        sectors_page = paginator.page(paginator.num_pages)

    context = {'sectors': sectors, 'sector_count': sector_count, 'sectors': sectors_page, 'paginator': paginator, 'page_obj': sectors_page, 'user': user}

    return render(request, 'sector.html', context)

# Service Page
def service(request):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = Staff.objects.get(id=staff_id)

    services = Service.objects.order_by('service_name')
    service_count = Service.objects.count()
    paginator = Paginator(services, 10)
    page = request.GET.get('page')

    try:
        services_page = paginator.page(page)
    except PageNotAnInteger:
        services_page = paginator.page(1)
    except EmptyPage:
        services_page = paginator.page(paginator.num_pages)

    context = {'services': services, 'service_count': service_count, 'services': services_page, 'paginator': paginator, 'page_obj': services_page, 'user': user}

    return render(request, 'service.html', context)

# Product
def product(request):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = Staff.objects.get(id=staff_id)

    products = Product.objects.order_by('product_name')
    product_count = Product.objects.count()
    paginator = Paginator(products, 10)
    page = request.GET.get('page')

    try:
        product_page = paginator.page(page)
    except PageNotAnInteger:
        product_page = paginator.page(1)
    except EmptyPage:
        product_page = paginator.page(paginator.num_pages)

    context = {'products': products, 'product_count': product_count, 'products': product_page, 'paginator': paginator, 'page_obj': product_page, 'user': user}

    return render(request, 'product.html', context)

# Brand Page
def brand(request):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = Staff.objects.get(id=staff_id)

    brands = Brand.objects.order_by('brand_name')
    brand_count = Brand.objects.count()
    paginator = Paginator(brands, 10)
    page = request.GET.get('page')

    try:
        brand_page = paginator.page(page)
    except PageNotAnInteger:
        brand_page = paginator.page(1)
    except EmptyPage:
        brand_page = paginator.page(paginator.num_pages)

    context = {'brands': brands, 'brand_count': brand_count, 'brands': brand_page, 'paginator': paginator, 'page_obj': brand_page, 'user': user}

    return render(request, 'brand.html', context)

def requirementform(request):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = get_object_or_404(Staff, id=staff_id)

    companies = Company.objects.all()
    contacts = Contact.objects.select_related('company').all()
    services = Service.objects.all()
    brands = Brand.objects.all()
    products = Product.objects.all()

    context = {'companies': companies, 'contacts': contacts, 'services': services, 'brands': brands, 'products': products, 'user': user}
    return render(request, 'requirementform.html', context)

# New Requirement Submission
def add_requirement(request):
    if request.method == 'POST':
        company_id = request.POST.get('company-name')
        contact_id = request.POST.get('contact-person')
        requirement_type = request.POST.get('requirement-category')
        brand_id = request.POST.get('brand') if requirement_type == 'Product' else None
        product_id = request.POST.get('product') if requirement_type == 'Product' else None
        service_id = request.POST.get('service') if requirement_type == 'Service' else None
        requirement_description = request.POST.get('message')
        currency = request.POST.get('currency')
        price = request.POST.get('price')
        status = request.POST.get('status') or 'Approved'
        progress = request.POST.get('progress')

        company = get_object_or_404(Company, pk=company_id)
        contact = Contact.objects.filter(pk=contact_id).first() if contact_id else None
        brand = Brand.objects.filter(pk=brand_id).first() if brand_id else None
        product = Product.objects.filter(pk=product_id).first() if product_id else None
        service = Service.objects.filter(pk=service_id).first() if service_id else None

        Requirement.objects.create(company=company, date=now(), contact_name=contact, requirement_type=requirement_type, brand=brand, product_name=product, service=service, requirement_description=requirement_description, currency=currency, price=price, status=status, progress=progress)

        return redirect(reverse('contract'))
    
# New Staff Submission
def add_newstaff(request):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_role = Staff.objects.filter(id=request.session['staff_id']).first()
    
    # Check if the current staff member exists and is an admin
    if not staff_role or staff_role.role.lower() != 'admin':
        return redirect('login')
        
    if request.method == 'POST':
        full_name = request.POST.get('staff-name')
        role = request.POST.get('role')
        email = request.POST.get('email-address')
        password = request.POST.get('password')
        
        if not full_name or not email or not password or not role:
            return redirect('staff')

        if Staff.objects.filter(email=email).exists():
            return redirect('staff')

        try:
            new_staff = Staff(full_name=full_name, role=role, email=email, password=make_password(password))
            new_staff.save()
            messages.success(request, "Staff member added successfully.")
            return redirect('staff')
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            return redirect('add_newstaff')

    return render(request, 'staff.html')

# New Minute Submission
def add_newtransaction(request):
    if 'staff_id' not in request.session:
        return redirect('login')

    if request.method == 'POST':
        company = get_object_or_404(Company, id=request.POST.get('company-name'))
        requirement = get_object_or_404(Requirement, id=request.POST.get('requirement'))
        contact = get_object_or_404(Contact, id=request.POST.get('contact-person'))

        Minute.objects.create(date=request.POST.get('date'), company=company, requirement=requirement, contact=contact, action=request.POST.get('action'), remark=request.POST.get('remarks'))
        return redirect('transaction')

# Add Minute
def add_transaction(request, company_id, requirement_id):
    if 'staff_id' not in request.session:
        return redirect('login')

    if request.method == 'POST':
        company = get_object_or_404(Company, id=company_id)
        requirement = get_object_or_404(Requirement, id=requirement_id)
        date = request.POST.get('date')
        contact_id = request.POST.get('contact-person')
        contact = get_object_or_404(Contact, id=contact_id)
        action = request.POST.get('action')
        remark = request.POST.get('remarks')

        Minute.objects.create(date=date, company=company, requirement=requirement, contact=contact, action=action, remark=remark )
        return redirect('contractdetails', company_id=company_id, requirement_id=requirement_id)
    
# New Sector Submission
def add_newsector(request):
    if 'staff_id' not in request.session:
        return redirect('login')

    if request.method == 'POST':
        sector_name = request.POST.get('sector-name')

        if Sector.objects.filter(sector_name=sector_name).exists():
            return redirect(reverse('sector'))

        sector = Sector(sector_name=sector_name)
        sector.save()

        return redirect(reverse('sector'))
    else:
        return render(request, 'sector.html')    

# New Service Submission
def add_newservice(request):
    if 'staff_id' not in request.session:
        return redirect('login')

    if request.method == 'POST':
        service_name = request.POST.get('service-name')

        if Service.objects.filter(service_name=service_name).exists():
            return redirect(reverse('service'))

        service = Service(service_name=service_name)
        service.save()

        return redirect(reverse('service'))
    else:
        return render(request, 'service.html')
    
# New Product Submission
def add_newproduct(request):
    if 'staff_id' not in request.session:
        return redirect('login')

    if request.method == 'POST':
        product_name = request.POST.get('product-name')

        if Product.objects.filter(product_name=product_name).exists():
            return redirect(reverse('product'))

        product = Product(product_name=product_name)
        product.save()

        return redirect(reverse('product'))
    else:
        return render(request, 'product.html')
    
# New Brand Submission
def add_newbrand(request):
    if 'staff_id' not in request.session:
        return redirect('login')

    if request.method == 'POST':
        brand_name = request.POST.get('brand-name')

        if Brand.objects.filter(brand_name=brand_name).exists():
            return redirect(reverse('brand'))

        brand = Brand(brand_name=brand_name)
        brand.save()

        return redirect('brand')
    else:
        return render(request, 'brand.html')
    
# Staff Status Toggle
@require_POST
def toggle_staff_status(request, staff_id):
    staff = get_object_or_404(Staff, id=staff_id)
    staff.status = not staff.status
    staff.save()
    return JsonResponse({'status': staff.status})

# View Company Details
def companydetails(request, company_id):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = get_object_or_404(Staff, id=staff_id)

    company = get_object_or_404(Company, id=company_id)
    requirements = Requirement.objects.filter(company=company).select_related('brand', 'product_name', 'service').prefetch_related('requirement_transactions')
    contacts = Contact.objects.filter(company=company)
    services = Service.objects.values('id', 'service_name').distinct().order_by('service_name')
    brands = Brand.objects.values('id', 'brand_name').distinct().order_by('brand_name')
    products = Product.objects.values('id', 'product_name').distinct().order_by('product_name')

    requirements_count = Requirement.objects.filter(company=company).select_related('brand', 'product_name', 'service').prefetch_related('requirement_transactions')
    requirements_paginator = Paginator(requirements_count, 10)
    requirements_page_number = request.GET.get('requirements_page', 1)
    requirements_page_obj = requirements_paginator.get_page(requirements_page_number)

    context = {'user': user, 'company': company, 'requirements': requirements, 'contacts': contacts, 'services': services, 'brands': brands, 'products': products, 'requirements_page_obj': requirements_page_obj}
    return render(request, 'companydetails.html', context)

# View Contract Details
def contractdetails(request, company_id, requirement_id):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = get_object_or_404(Staff, id=staff_id)
    
    company = get_object_or_404(Company, id=company_id)
    requirement = get_object_or_404(Requirement, id=requirement_id)
    requirements = Requirement.objects.filter(company=company).select_related('brand', 'product_name', 'service').prefetch_related('requirement_transactions')
    transactions = Minute.objects.filter(company=company, requirement=requirement).order_by('-date')

    transactions_count = Minute.objects.filter(company=company, requirement=requirement).select_related('date', 'action', 'remark').prefetch_related('transaction_transactions')
    transactions_paginator = Paginator(transactions_count, 10)
    transactions_page_number = request.GET.get('transactions_page', 1)
    transactions_page_obj = transactions_paginator.get_page(transactions_page_number)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date and end_date:
        transactions = transactions.filter(date__range=[start_date, end_date])

    contacts = Contact.objects.filter(company=company)
    context = {'user': user, 'company': company, 'requirement': requirement, 'requirements': requirements, 'transactions': transactions, 'transactions_page_obj': transactions_page_obj, 'contacts': contacts}
    return render(request, 'contractdetails.html', context)

# View Task Details
def taskdetails(request, task_id):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = get_object_or_404(Staff, id=staff_id)
    task = get_object_or_404(Task, id=task_id)

    context = {'user': user, 'task': task}
    return render(request, 'taskdetails.html', context)

# Edit Company Details
def companyeditform(request, company_id):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    COUNTRY_CHOICES = [
        ('Nepal', 'Nepal'),
        ('India', 'India'),
        ('Singapore', 'Singapore'),
    ]

    ROLES = [
        ('Client', 'Client'),
        ('Partner', 'Partner')
    ]

    partner_companies = Company.objects.filter(role='Partner').order_by('company_name')
    
    company = Company.objects.get(pk=company_id)
    contacts = Contact.objects.filter(company=company)
    sectors = Sector.objects.all()

    religions = ["Hinduism", "Buddhism", "Christianity"]
    return render(request, 'companyeditform.html', {'country_choices': COUNTRY_CHOICES, 'roles': ROLES, 'company': company, 'contacts': contacts, 'sectors': sectors, 'partner_companies': partner_companies, 'religions': religions})

# Edit Requirement Details
def requirementeditform(request, requirement_id):
    if 'staff_id' not in request.session:
        return redirect('login')
    
    staff_id = request.session.get('staff_id')
    user = get_object_or_404(Staff, id=staff_id)

    requirement = get_object_or_404(Requirement, id=requirement_id)
    company = requirement.company
    contacts = Contact.objects.filter(company=company)
    services = Service.objects.values('id', 'service_name').distinct().order_by('service_name')
    brands = Brand.objects.values('id', 'brand_name').distinct().order_by('brand_name')
    products = Product.objects.values('id', 'product_name').distinct().order_by('product_name')

    return render(request, 'requirementeditform.html', {'user': user, 'requirement': requirement, 'contacts': contacts, 'services': services, 'brands': brands, 'products': products})

# Update Company
def update_company(request, company_id):
    company = get_object_or_404(Company, id=company_id)

    if request.method == 'POST':
        company.company_name = request.POST['company_name']
        company.role = request.POST['role']
        company.sector_id = request.POST['sector']
        company.address = request.POST['address']
        company.city = request.POST['city']
        company.state = request.POST['state']
        company.country = request.POST['country']
        company.via = request.POST['via']
        company.referral_name = request.POST.get('referral_name', '')
        partner_id = request.POST.get('partner', None)
        if partner_id:
            try:
                company.partner_company = Company.objects.get(id=partner_id)
            except Company.DoesNotExist:
                company.partner_company = None
        else:
            company.partner_company = None
            
        company.website = request.POST['website']
        company.connection = request.POST['connection']
        company.save()

        contact_ids = request.POST.getlist('contact_id[]')
        contact_names = request.POST.getlist('contact_name[]')
        designations = request.POST.getlist('designation[]')
        emails = request.POST.getlist('email[]')
        phone_numbers = request.POST.getlist('phone_number[]')
        dobs = request.POST.getlist('dob[]')
        religions = request.POST.getlist('religion[]')

        existing_contacts = set(Contact.objects.filter(company=company).values_list('id', flat=True))
        updated_contacts = set([int(contact_id) for contact_id in contact_ids if contact_id])
        contacts_to_delete = existing_contacts - updated_contacts
        Contact.objects.filter(id__in=contacts_to_delete).delete()

        for i in range(len(contact_names)):
            contact_id = contact_ids[i] if contact_ids[i] else None
            if contact_id:
                contact = Contact.objects.get(id=contact_id, company=company)
            else:
                contact = Contact(company=company)
            contact.contact_name = contact_names[i]
            contact.designation = designations[i]
            contact.email = emails[i]
            contact.phone_number = phone_numbers[i]
            contact.dob = dobs[i] or None
            contact.religion = religions[i]
            contact.save()

        return redirect('companyeditform', company_id=company_id)

    sectors = Sector.objects.all()
    contacts = Contact.objects.filter(company=company)

    return render(request, 'company.html', {'company': company, 'sectors': sectors, 'contacts': contacts})

# Update Task Status
def update_task_status(request, task_id):
    if 'staff_id' not in request.session:
        return redirect('login')

    # Get the task to be updated
    task = get_object_or_404(Task, id=task_id)

    # Check if the request method is POST
    if request.method == 'POST':
        # Prevent status change if task is already completed
        if task.status == 'Completed':
            messages.error(request, "This task has already been completed and cannot be updated.")
            return redirect('taskdetails', task_id=task.id)

        # Get the new status from the form
        new_status = request.POST.get('status')

        # Update the task's status if it's not completed
        task.status = new_status
        task.save()

        # Redirect to the task details page after update
        messages.success(request, "Task status updated successfully.")
        return redirect('taskdetails', task_id=task.id)

    return HttpResponse("Invalid request")

# Add New Requirement
def add_newrequirement(request):
    if request.method == 'POST':
        company_id = request.POST.get('company-name')
        contact_id = request.POST.get('contact-person')
        requirement_type = request.POST.get('requirement-category')
        brand_id = request.POST.get('brand') if requirement_type == 'Product' else None
        product_id = request.POST.get('product') if requirement_type == 'Product' else None
        service_id = request.POST.get('service') if requirement_type == 'Service' else None
        requirement_description = request.POST.get('message')
        currency = request.POST.get('currency')
        price = request.POST.get('price')
        status = request.POST.get('status') or 'Approved'
        progress = request.POST.get('progress')

        company = get_object_or_404(Company, pk=company_id)
        contact = Contact.objects.filter(pk=contact_id).first() if contact_id else None
        brand = Brand.objects.filter(pk=brand_id).first() if brand_id else None
        product = Product.objects.filter(pk=product_id).first() if product_id else None
        service = Service.objects.filter(pk=service_id).first() if service_id else None

        Requirement.objects.create(company=company, date=now(), contact_name=contact, requirement_type=requirement_type, brand=brand, product_name=product, service=service, requirement_description=requirement_description, currency=currency, price=price, status=status, progress=progress)

        return redirect('companydetails', company_id=company.id)

# Update Requirement
def update_requirement(request, requirement_id):
    requirement = get_object_or_404(Requirement, id=requirement_id)

    if request.method == 'POST':
        contact_id = request.POST['contact-person']
        contact = get_object_or_404(Contact, id=contact_id)
        requirement.contact_name = contact

        requirement.date = request.POST['date']
        requirement.requirement_type = request.POST['requirement-category']
        requirement.service = get_object_or_404(Service, id=request.POST['service'])
        requirement.brand = get_object_or_404(Brand, id=request.POST['brand'])
        requirement.product_name = get_object_or_404(Product, id=request.POST['product'])
        requirement.requirement_description = request.POST['message']
        requirement.currency = request.POST['currency']
        requirement.price = request.POST['price']
        requirement.progress = request.POST['progress']
        requirement.save()

        return redirect('requirementeditform', requirement_id=requirement.id)

    return redirect('requirementeditform', requirement_id=requirement.id)

# Update Sector
@require_POST
def update_sector(request):
    sector_id = request.POST.get('sector-id')
    sector_name = request.POST.get('sector-name')
    sector = get_object_or_404(Sector, id=sector_id)

    sector.sector_name = sector_name
    sector.save()

    return redirect('sector')

# Update Service
@require_POST
def update_service(request):
    service_id = request.POST.get('service-id')
    service_name = request.POST.get('service-name')
    service = get_object_or_404(Service, id=service_id)

    service.service_name = service_name
    service.save()

    return redirect('service')

# Update Product
@require_POST
def update_product(request):
    product_id = request.POST.get('product-id')
    product_name = request.POST.get('product-name')
    product = get_object_or_404(Product, id=product_id)

    product.product_name = product_name
    product.save()

    return redirect('product')

# Update Brand
@require_POST
def update_brand(request):
    brand_id = request.POST.get('brand-id')
    brand_name = request.POST.get('brand-name')
    brand = get_object_or_404(Brand, id=brand_id)

    brand.brand_name = brand_name
    brand.save()

    return redirect('brand')

# Get AJAX Contacts for Minute Form
def get_contacts(request):
    company_id = request.GET.get('company_id')
    contacts = Contact.objects.filter(company_id=company_id).values('id', 'contact_name')
    return JsonResponse(list(contacts), safe=False)

def get_requirements(request):
    company_id = request.GET.get('company_id')
    requirements = Requirement.objects.filter(company_id=company_id)
    data = []

    for requirement in requirements:
        if requirement.requirement_type == 'Product':
            display_text = f"{requirement.brand.brand_name if requirement.brand else ''} - {requirement.product_name.product_name if requirement.product_name else ''}"
        elif requirement.requirement_type == 'Service':
            display_text = requirement.service.service_name if requirement.service else 'No Service Name'
        else:
            display_text = 'Unknown Requirement'

        data.append({'id': requirement.id, 'requirement_name': display_text})
    
    return JsonResponse(data, safe=False)

# Export Excel File
def export_excel(request, company_id):
    if 'staff_id' not in request.session:
        return redirect('login')

    try:
        company = Company.objects.get(pk=company_id)
    except Company.DoesNotExist:
        return HttpResponse("Company not found.", status=404)

    transactions = Minute.objects.filter(company_id=company_id).select_related('requirement').order_by('requirement__id', '-date')

    wb = Workbook()
    ws = wb.active
    ws.title = "Company Data"

    # Define styles
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Add headers for contacts
    headers = ['Company Name', 'Contact Person', 'Email', 'Phone Number']
    ws.append(headers)
    for cell in ws[1]:
        cell.alignment = center_alignment
        cell.font = bold_font
        cell.border = thin_border

    # Add company contact persons
    for contact in company.company_contacts.all():
        ws.append([company.company_name, contact.contact_name, contact.email, contact.phone_number])
        for cell in ws[ws.max_row]:
            cell.border = thin_border

    # Leave a blank row
    ws.append([])
    ws.append(['S.N.', 'Requirement Type', 'Requirement', 'Date', 'Action', 'Remark'])

    for cell in ws[ws.max_row]:
        cell.alignment = center_alignment
        cell.font = bold_font
        cell.border = thin_border

    # Set column widths
    column_widths = {'A': 30, 'B': 30, 'C': 35, 'D': 20, 'E': 50, 'F': 50}
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    # Write transaction data
    last_requirement_id = None
    merge_start_row = ws.max_row + 1
    sn = 1

    for transaction in transactions:
        requirement = transaction.requirement
        requirement_id = requirement.id if requirement else None
        requirement_type = requirement.requirement_type if requirement else ''
        requirement_value = ''

        if requirement_type.lower() == 'product':
            brand = requirement.brand.brand_name if requirement and requirement.brand else ''
            product_name = requirement.product_name.product_name if requirement and requirement.product_name else ''
            if brand and product_name:
                requirement_value = f"{brand} - {product_name}"
        elif requirement_type.lower() == 'service':
            requirement_value = requirement.service.service_name if requirement and requirement.service else ''

        if last_requirement_id and last_requirement_id != requirement_id:
            if ws.max_row - 1 > merge_start_row:
                ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=ws.max_row - 1, end_column=1)
                ws.merge_cells(start_row=merge_start_row, start_column=2, end_row=ws.max_row - 1, end_column=2)
                ws.merge_cells(start_row=merge_start_row, start_column=3, end_row=ws.max_row - 1, end_column=3)
            merge_start_row = ws.max_row + 1

        ws.append([sn, requirement_type, requirement_value, transaction.date, transaction.action, transaction.remark])
        for cell in ws[ws.max_row]:
            cell.alignment = left_alignment if cell.column_letter in ['E', 'F'] else center_alignment
            cell.border = thin_border

        last_requirement_id = requirement_id
        sn += 1

    if ws.max_row >= merge_start_row:
        ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=ws.max_row, end_column=1)
        ws.merge_cells(start_row=merge_start_row, start_column=2, end_row=ws.max_row, end_column=2)
        ws.merge_cells(start_row=merge_start_row, start_column=3, end_row=ws.max_row, end_column=3)

    # Merge cells for company name
    company_name_start_row = 2
    company_name_end_row = company.company_contacts.count() + 1
    ws.merge_cells(start_row=company_name_start_row, start_column=1, end_row=company_name_end_row, end_column=1)
    for row in range(company_name_start_row, company_name_end_row + 1):
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].border = thin_border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{company.company_name} - Report.xlsx"'
    return response

def requirement_list_view(request):
    requirements = Requirement.objects.all()
    return render(request, 'requirement.html', {'requirements': requirements})

# AJAX
@csrf_exempt
def add_sector_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            sector_name = data.get('sector_name')

            if not sector_name:
                return JsonResponse({'error': 'Sector name is required'}, status=400)

            if Sector.objects.filter(sector_name=sector_name).exists():
                return JsonResponse({'error': 'Sector already exists'}, status=400)

            sector = Sector.objects.create(sector_name=sector_name)
            return JsonResponse({'id': sector.id, 'sector_name': sector.sector_name}, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=405)
    
@csrf_exempt
def add_service_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            service_name = data.get('service_name')

            if not service_name:
                return JsonResponse({'error': 'Service name is required'}, status=400)

            if Service.objects.filter(service_name=service_name).exists():
                return JsonResponse({'error': 'Service already exists'}, status=400)

            service = Service.objects.create(service_name=service_name)
            return JsonResponse({'id': service.id, 'service_name': service.service_name}, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=405)
    
@csrf_exempt
def add_brand_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            brand_name = data.get('brand_name')

            if not brand_name:
                return JsonResponse({'error': 'Brand name is required'}, status=400)

            if Brand.objects.filter(brand_name=brand_name).exists():
                return JsonResponse({'error': 'Brand already exists'}, status=400)

            brand = Brand.objects.create(brand_name=brand_name)
            return JsonResponse({'id': brand.id, 'brand_name': brand.brand_name}, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=405)
    
@csrf_exempt
def add_product_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            product_name = data.get('product_name')

            if not product_name:
                return JsonResponse({'error': 'Product name is required'}, status=400)

            if Product.objects.filter(product_name=product_name).exists():
                return JsonResponse({'error': 'Product already exists'}, status=400)

            product = Product.objects.create(product_name=product_name)
            return JsonResponse({'id': product.id, 'product_name': product.product_name}, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=405)

# Client Forgot Password
def forget_pass(request):
    return render(request, "forget_pass.html")

def reset_code(request):
    message = ""
    if request.method == "POST":
        email = request.POST['email']
        message = None
        print(email)  # in future we will send the code to the provided if it exist
        # generate code and send via email
        if User.objects.filter(email=email).exists():
            random_float = random.randint(100000, 999999)
            print(random_float)
            c_details = User.objects.get(email=email)

            # saving the generated password reset code to database
            c_details.resetcode = random_float
            c_details.save()

            # creating the session to send email of user to change password in reset_passwordDone method
            request.session['customer_email'] = c_details.email

            #sending forget password code to user mail

            user = User.objects.get(email = email)

            subject = "Reset Password"
            html_content = render_to_string('email.html',{
                                            'fname': user.full_name, 'email': user.email,'code':random_float})
            from_email = 'zaynierashik@gmail.com'
            to = [c_details.email]

            text_content = strip_tags(html_content)
            email = EmailMultiAlternatives(
                subject,
                text_content,
                from_email,
                to,
            )
            email.attach_alternative(html_content, "text/html")
            email.send(fail_silently=False)

            return render(request, "code_reset.html", {'fname': c_details.full_name,'code':c_details.resetcode})
        else:
            message = "This email doesn't exist."
            print("This email doesn't exist.")
    return render(request, "forget_pass.html",{
        "toast_message" : message,
    })


def reset_password(request):
    email = request.session.get('customer_email')
    print(email)
    customer_detail = User.objects.get(email=email)
    message = None  # Initialize the message variable

    if request.method == "POST":
        entered_email = request.POST['email']
        entered_code = request.POST['code']
        print(customer_detail.email)

        # Validate the provided email and code
        if customer_detail.email != entered_email and int(customer_detail.resetcode) != int(entered_code):
            message = "Both Email and Code are incorrect."
            print("Both Email and Code are incorrect.")
        elif customer_detail.email != entered_email:
            message = "Invalid Email provided."
            print("Invalid Email provided.")
        elif int(customer_detail.resetcode) != int(entered_code):
            message = "Invalid Code provided."
            print("Invalid Code provided.")
        else:
            customer_detail.resetcode = None
            customer_detail.save()
            return render(request, "reset_password.html")

    return render(request, "code_reset.html", {
        "toast_message": message,
    })


def reset_passwordDone(request):
    message = None
    if request.method == "POST":
        password = request.POST['password']
        c_password = request.POST['c_password']
        email = request.session.get('customer_email')
        # Access the user data via email
        customer_detail = User.objects.get(email=email)

        if password != c_password:
            message = "Confirm password didn't match."
            print("Confirm password didn't match.")
        elif check_password(password, customer_detail.password):
            message = "Password can't be same as the old one."
            print("Password can't be same as the old one.")
        else:
            # Save the new password
            customer_detail.password = make_password(password)
            customer_detail.save()
            print("Password Updated.")
            return render(request, "userauthentication.html")

    return render(request, "reset_password.html", {
        "toast_message": message,
    })


# Admin Forgot Password
def admin_forget_pass(request):
    return render(request, "forget_pass copy.html")


def admin_reset_code(request):
    message = ""
    if request.method == "POST":
        email = request.POST['email']
        message = None
        print(email)  # in future we will send the code to the provided if it exist
        # generate code and send via email
        if Staff.objects.filter(email=email).exists():
            random_float = random.randint(100000, 999999)
            print(random_float)
            c_details = Staff.objects.get(email=email)

            # saving the generated password reset code to database
            c_details.resetcode = random_float
            c_details.save()

            # creating the session to send email of user to change password in reset_passwordDone method
            request.session['customer_email'] = c_details.email

            #sending forget password code to user mail

            user = Staff.objects.get(email = email)

            subject = "Reset Password"
            html_content = render_to_string('email.html',{
                                            'fname': user.full_name, 'email': user.email,'code':random_float})
            from_email = 'zaynierashik@gmail.com'
            to = [c_details.email]

            text_content = strip_tags(html_content)
            email = EmailMultiAlternatives(
                subject,
                text_content,
                from_email,
                to,
            )
            email.attach_alternative(html_content, "text/html")
            email.send(fail_silently=False)

            return render(request, "code_reset copy.html", {'fname': c_details.full_name,'code':c_details.resetcode})
        else:
            message = "This email doesn't exist."
            print("This email doesn't exist.")
    return render(request, "forget_pass copy.html",{
        "toast_message" : message,
    })


def admin_reset_password(request):
    email = request.session.get('customer_email')
    print(email)
    customer_detail = Staff.objects.get(email=email)
    message = None  # Initialize the message variable

    if request.method == "POST":
        entered_email = request.POST['email']
        entered_code = request.POST['code']
        print(customer_detail.email)

        # Validate the provided email and code
        if customer_detail.email != entered_email and int(customer_detail.resetcode) != int(entered_code):
            message = "Both Email and Code are incorrect."
            print("Both Email and Code are incorrect.")
        elif customer_detail.email != entered_email:
            message = "Invalid Email provided."
            print("Invalid Email provided.")
        elif int(customer_detail.resetcode) != int(entered_code):
            message = "Invalid Code provided."
            print("Invalid Code provided.")
        else:
            customer_detail.resetcode = None
            customer_detail.save()
            return render(request, "reset_password copy.html")

    return render(request, "code_reset copy.html", {
        "toast_message": message,
    })


def admin_reset_passwordDone(request):
    message = None
    if request.method == "POST":
        password = request.POST['password']
        c_password = request.POST['c_password']
        email = request.session.get('customer_email')
        # Access the user data via email
        customer_detail = Staff.objects.get(email=email)

        if password != c_password:
            message = "Confirm password didn't match."
            print("Confirm password didn't match.")
        elif check_password(password, customer_detail.password):
            message = "Password can't be same as the old one."
            print("Password can't be same as the old one.")
        else:
            # Save the new password
            customer_detail.password = make_password(password)
            customer_detail.save()
            print("Password Updated.")
            return render(request, "login.html")

    return render(request, "reset_password copy.html", {
        "toast_message": message,
    })

# Get states
def get_states(request):
    country = request.GET.get('country')
    
    states = {
        'Nepal': ['Province 1', 'Madhesh Pradesh', 'Bagmati', 'Gandaki', 'Lumbini', 'Karnali', 'Sudurpashchim'],
        'India': [
            'Andhra Pradesh', 'Arunachal Pradesh', 'Assam', 'Bihar', 'Chhattisgarh', 'Goa', 'Gujarat',
            'Haryana', 'Himachal Pradesh', 'Jharkhand', 'Karnataka', 'Kerala', 'Madhya Pradesh', 'Maharashtra',
            'Manipur', 'Meghalaya', 'Mizoram', 'Nagaland', 'Odisha', 'Punjab', 'Rajasthan', 'Sikkim',
            'Tamil Nadu', 'Telangana', 'Tripura', 'Uttar Pradesh', 'Uttarakhand', 'West Bengal', 
            'Andaman and Nicobar Islands', 'Chandigarh', 'Dadra and Nagar Haveli and Daman and Diu',
            'Lakshadweep', 'Delhi', 'Puducherry'
        ],
        'Singapore': ['Central Region', 'East Region', 'North Region', 'North-East Region', 'West Region']
    }

    # Return an empty list if country is not in the predefined countries
    states_list = states.get(country, [])

    return JsonResponse({'states': states_list})