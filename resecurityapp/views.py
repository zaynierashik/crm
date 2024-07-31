import io
import pandas as pd
from django.shortcuts import *
from django.http import *
from .models import *
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import *
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from django.contrib.auth.hashers import check_password, make_password
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.utils import *

def submit_signup(request):
    if request.method == 'POST':
        fullname = request.POST.get('fullname')
        email = request.POST.get('email')
        password = request.POST.get('password')
        hashed_password = make_password(password)
        role = request.POST.get('role-type')

        name_exists = Staff.objects.filter(Full_Name=fullname).exists()
        email_exists = Staff.objects.filter(email=email).exists()

        if name_exists and email_exists:
            redirect_url = reverse('master') + '?selection=staff&success=False&message=Both name and email already exist.'
        elif name_exists:
            redirect_url = reverse('master') + '?selection=staff&success=False&message=Staff with this name already exists.'
        elif email_exists:
            redirect_url = reverse('master') + '?selection=staff&success=False&message=Email already exists.'
        else:
            user = Staff(Full_Name=fullname, email=email, password=hashed_password, role=role)
            user.save()
            redirect_url = reverse('master') + '?selection=staff&success=True&message=Staff account created successfully.'

        return redirect(redirect_url)
    else:
        return redirect(reverse('master') + '?selection=staff')

def login(request):
    fullname = request.session.get('fullname')

    if fullname:
        return redirect(reverse('master') + '?selection=company')
    
    return render(request, 'login.html')

def submit_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        try:
            ss = Staff.objects.get(email=email)
            if not ss.status:
                return redirect(reverse('login') + '?success=False&reason=inactive')
            
            if check_password(password, ss.password):
                request.session['fullname'] = ss.Full_Name
                request.session['staffrole'] = ss.role
                return redirect(reverse('master') + '?selection=company')
            else:
                return redirect(reverse('login') + '?success=False&reason=wrong_password')
        except Staff.DoesNotExist:
            return redirect(reverse('login') + '?success=False&reason=not_found')

    return render(request, 'login.html')

@require_POST
def toggle_staff_status(request, staff_id):
    staff = get_object_or_404(Staff, id=staff_id)
    staff.status = not staff.status
    staff.save()
    return JsonResponse({'status': staff.status})

def logout(request):
    request.session.flush()
    return redirect('login')   

def transaction(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    companies = Company.objects.filter(status='active', Created_By=fullname).order_by('Company_Name')
    contacts = Contact.objects.all()
    requirements = Requirement.objects.all()
    services = Service.objects.all()
    brands = Brand.objects.all()
    products = Product.objects.all()
    success = request.GET.get('success')
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
        
    return render(request, 'transaction.html', {'companies': companies, 'contacts': contacts, 'requirements': requirements, 'services': services, 'brands': brands, 'products': products, 'success': success, 'fullname': fullname})

def report(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    companies = Company.objects.filter(status='active', Created_By=fullname).order_by('Company_Name')
    contacts = Contact.objects.all()
    requirements = Requirement.objects.all()
    services = Service.objects.all()
    brands = Brand.objects.all()
    products = Product.objects.all()
        
    return render(request, 'report.html', {'companies': companies, 'contacts': contacts, 'requirements': requirements, 'services': services, 'brands': brands, 'products': products})

def generate_report(request):
    if request.method == 'POST':
        company_id = request.POST.get('company')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        if company_id:
            companies = Company.objects.filter(pk=company_id)
        else:
            companies = Company.objects.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Company Report"
        ws.append(['Date', 'Company Name', 'Address', 'City', 'State', 'Country', 'Contact Person(s)', 'Email', 'Brand', 'Product Name', 'Service', 'Price', 'Status', 'Action', 'Remark'])

        column_widths = {'A': 15, 'B': 17, 'C': 25, 'D': 25, 'E': 35, 'F': 35, 'G': 35, 'H': 25, 'I': 10, 'J': 35, 'K': 25, 'L': 25, 'M': 25, 'N': 25, 'O': 25}
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

        date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')
        center_alignment = Alignment(horizontal='center', vertical='center')
        wrap_text_alignment = Alignment(wrap_text=True, vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.alignment = center_alignment
            cell.border = thin_border

        for company in companies:
            transactions = Transaction.objects.filter(company=company).order_by('-date')

            if start_date and end_date:
                transactions = transactions.filter(date__range=[start_date, end_date])

            company_row_start = ws.max_row + 1
            contact_names = set()
            for transaction in transactions:
                brand = transaction.requirement.brand.Brand_Name if transaction.requirement and transaction.requirement.brand else ''
                service = transaction.requirement.service.Service_Name if transaction.requirement and transaction.requirement.service else ''
                contact_person = transaction.contact.Contact_Name if transaction.contact else ''
                contact_email = transaction.contact.email if transaction.contact else ''
                price = transaction.requirement.price if transaction.requirement else ''
                status = transaction.requirement.status if transaction.requirement else ''
                product_name = transaction.requirement.Product_Name if transaction.requirement and hasattr(transaction.requirement, 'Product_Name') else ''

                if contact_person:
                    contact_names.add(contact_person)

                ws.append([
                    str(transaction.date),
                    str(transaction.company.Company_Name),
                    str(transaction.company.address),
                    str(transaction.company.city),
                    str(transaction.company.state),
                    str(transaction.company.country),
                    '',
                    str(contact_email),
                    str(brand),
                    str(product_name),
                    str(service),
                    str(price),
                    str(status),
                    str(transaction.action),
                    str(transaction.remark)
                ])

                row_num = ws.max_row
                for col in ['H', 'I', 'J', 'K', 'O']:
                    ws[f'{get_column_letter(ws.max_column - 4)}{row_num}'].alignment = wrap_text_alignment

                for col in range(1, ws.max_column + 1):
                    ws[f'{get_column_letter(col)}{row_num}'].border = thin_border

            company_row_end = ws.max_row

            if company_row_end > company_row_start:
                for col in range(2, 7):  # Columns B to F (Company Name, Address, City, State, Country)
                    ws.merge_cells(start_row=company_row_start, start_column=col, end_row=company_row_end, end_column=col)

                # Merge and write contact names
                ws.merge_cells(start_row=company_row_start, start_column=7, end_row=company_row_end, end_column=7)
                ws.cell(row=company_row_start, column=7, value=', '.join(contact_names))
                ws.cell(row=company_row_start, column=7).alignment = wrap_text_alignment

            if transactions.exists():
                ws.append([''] * 15)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Company_Report.xlsx"'
        wb.save(response)

        return response

    companies = Company.objects.all()
    return render(request, 'report.html', {'companies': companies})
    

def get_companydetails(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    company_name = request.GET.get('company')
    requirement_type = request.GET.get('requirement_type')
    brand_name = request.GET.get('brand')
    company = get_object_or_404(Company, Company_Name=company_name)

    if requirement_type == 'Product':
        if brand_name:
            products = list(Requirement.objects.filter(company=company, Requirement_Type='Product', brand__Brand_Name=brand_name).values('Product_Name'))
        else:
            products = list(Requirement.objects.filter(company=company, Requirement_Type='Product').values('Product_Name', 'brand__Brand_Name'))
        return JsonResponse({'products': products})
    elif requirement_type == 'Service':
        services = list(Requirement.objects.filter(company=company, Requirement_Type='Service').values('service__Service_Name'))
        return JsonResponse({'services': services})

    return JsonResponse({'error': 'Invalid requirement type'}, status=400)

def submit_transaction(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        date = request.POST.get('date')
        company_name = request.POST.get('company')
        requirement_id = request.POST.get('requirement')
        contact_name = request.POST.get('contact-name')
        action = request.POST.get('action')
        remark = request.POST.get('remark')
        contact_name = request.POST.get('contact-name')

        try:
            company = Company.objects.get(Company_Name=company_name)
        except Company.DoesNotExist:
            return HttpResponseBadRequest("Company does not exist.")
        
        try:
            contact = Contact.objects.get(Contact_Name=contact_name)
        except Contact.DoesNotExist:
            return HttpResponseBadRequest("Contact does not exist.")
        
        try:
            requirement = Requirement.objects.get(id=requirement_id)
        except Requirement.DoesNotExist:
            return HttpResponseBadRequest("Requirement does not exist.")

        transaction = Transaction(date=date, company=company, requirement=requirement, contact=contact, action=action, remark=remark)
        transaction.save()
        return redirect(reverse('transaction') + '?success=True')
    else:
        return HttpResponse("Form Submission Error!")

def transactiondetails(request, company_id, requirement_id):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    company = get_object_or_404(Company, id=company_id)
    requirement = get_object_or_404(Requirement, id=requirement_id)
    transactions = Transaction.objects.filter(company=company, requirement=requirement).order_by('-date')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date and end_date:
        transactions = transactions.filter(date__range=[start_date, end_date])

    contacts = Contact.objects.filter(company=company)
    context = {'company': company, 'requirement': requirement, 'transactions': transactions, 'contacts': contacts}
    return render(request, 'transactiondetails.html', context)

def master(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    companies = Company.objects.filter(Created_By=fullname).order_by('Company_Name')
    contacts = Contact.objects.select_related('company').all()
    sectors = Sector.objects.values('Sector_Name').order_by('Sector_Name')
    services = Service.objects.values('id', 'Service_Name').distinct().order_by('Service_Name')
    brands = Brand.objects.values('Brand_Name').distinct().order_by('Brand_Name')
    products = Product.objects.values('Product_Name').distinct().order_by('Product_Name')
    partners = Partner.objects.all().order_by('Partner_Name')
    cities = Company.objects.filter(Created_By=fullname).values_list('city', flat=True).distinct().order_by('city')
    staffs = Staff.objects.all()

    city_filter = request.GET.get('city', None)
    status_filter = request.GET.get('status', 'active')

    if city_filter:
        companies = companies.filter(city=city_filter)

    if status_filter:
        companies = companies.filter(status=status_filter)

    selection = request.GET.get('selection', None)
    return render(request, 'master.html', {'companies': companies, 'contacts': contacts, 'sectors': sectors, 'services': services, 'brands': brands, 'products': products, 'partners': partners, 'cities': cities, 'staffs': staffs, 'selection': selection, 'fullname': fullname})

def get_requirements(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    company_name = request.GET.get('company', None)
    requirements = Requirement.objects.filter(company__Company_Name=company_name).select_related('brand', 'Product_Name', 'service').values('id', 'Requirement_Type', 'Product_Name__Product_Name', 'service__Service_Name', 'brand__Brand_Name')

    return JsonResponse({'requirements': list(requirements)})

def get_contacts(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    company_name = request.GET.get('company', None)
    contacts = Contact.objects.filter(company__Company_Name=company_name).values('Contact_Name')
    return JsonResponse({'contacts': list(contacts)})

@csrf_exempt
def submit_contact(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        try:
            company_name = request.POST.get('company')
            contact_name = request.POST.get('contact_name')
            designation = request.POST.get('designation')
            email = request.POST.get('email')
            phone = request.POST.get('number')
            company = Company.objects.get(Company_Name=company_name)

            if Contact.objects.filter(company=company, Contact_Name=contact_name).exists():
                return JsonResponse({'error': 'A contact with this name already exists for the given company.'}, status=400)

            contact = Contact(company=company, Contact_Name=contact_name, designation=designation, email=email, Phone_Number=phone)
            contact.save()

            return JsonResponse({
                'Contact_Name': contact.Contact_Name,
                'company': company_name
            })
        except Company.DoesNotExist:
            return JsonResponse({'error': 'Company does not exist.'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Invalid request method.'}, status=400)
    
def newform(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    formtype = request.GET.get('formtype', None)
    return render(request, 'form.html', {'formtype': formtype})

def newcompany(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    sectors = Sector.objects.values('Sector_Name')
    services = Service.objects.values('Service_Name').distinct()
    brands = Brand.objects.values('Brand_Name').distinct()
    partners = Partner.objects.values('Partner_Name')
    
    return render(request, 'newcompany.html', {'sectors': sectors, 'services': services, 'brands': brands, 'partners': partners})

def submit_newcompany(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        company_name = request.POST.get('company')

        if Company.objects.filter(Company_Name=company_name).exists():
            return redirect(reverse('newcompany') + f'?error_message=Company with this name already exists!&company_name={company_name}')

        sector_name = request.POST.get('sector')
        sector = Sector.objects.filter(Sector_Name=sector_name)

        if sector.exists():
            sectors = sector.first()
        else:
            return HttpResponse("Sector not found")
        
        address = request.POST.get('address')
        city = request.POST.get('city')
        state = request.POST.get('state')
        country = request.POST.get('country')
        
        via_name = request.POST.get('via')
        website = request.POST.get('website')
        
        if via_name == 'Referral':
            referral_name = request.POST.get('referral_name')
            partner_name = None
        elif via_name == 'Partner':
            partner = request.POST.get('partner')
            partner_name = Partner.objects.filter(Partner_Name=partner).first()
            referral_name = None
        else:
            referral_name = None
            partner_name = None

        contact_names = request.POST.getlist('contact_name[]')
        designations = request.POST.getlist('designation[]')
        emails = request.POST.getlist('email[]')
        phone_numbers = request.POST.getlist('number[]')
        dobs = request.POST.getlist('dob[]')
        religions = request.POST.getlist('religion[]')

        company = Company(Company_Name=company_name, sector=sectors, address=address, city=city, state=state, country=country, via=via_name, Referral_Name=referral_name, Partner_Name=partner_name, website=website, Created_By=fullname)
        company.save()

        for i in range(len(contact_names)):
            if contact_names[i] and designations[i] and emails[i] and phone_numbers[i]:
                dob = dobs[i] if dobs[i] else None
                contact_person = Contact(company=company, Contact_Name=contact_names[i], designation=designations[i], email=emails[i], Phone_Number=phone_numbers[i], DOB=dob, religion=religions[i])
                contact_person.save()
                
        return redirect(reverse('master') + '?selection=company&success=True')
    else:
        return HttpResponse("Form Submission Error!")
    
def companyform(request, company_id):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    company = Company.objects.get(pk=company_id)
    contacts = Contact.objects.filter(company=company)
    sectors = Sector.objects.values('Sector_Name')
    partners = Partner.objects.all()

    display_via = ''
    if company.via and company.via == 'Direct':
        display_via = 'Direct'
    elif company.via and company.via == 'Referral' and company.Referral_Name:
        display_via = company.Referral_Name
    elif company.via and company.via == 'Partner' and company.Partner_Name:
        display_via = company.Partner_Name.Partner_Name

    countries = ["Nepal", "USA", "India", "Singapore"]
    religions = ["Hinduism", "Buddhism", "Christianity"]
    return render(request, 'companyform.html', {'company': company, 'contacts': contacts, 'sectors': sectors, 'partners': partners, 'countries': countries, 'religions': religions, 'display_via': display_via})

def submit_company(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        company_id = request.POST.get('company_id')
        company = get_object_or_404(Company, id=company_id)
        
        company.Company_Name = request.POST.get('company')
        sector_name = request.POST.get('sector')
        company.sector = get_object_or_404(Sector, Sector_Name=sector_name)
        company.address = request.POST.get('address')
        company.city = request.POST.get('city')
        company.state = request.POST.get('state')
        company.country = request.POST.get('country')
        company.website = request.POST.get('website')
        company.status = request.POST.get('status')

        company.save()
        company.contact_persons.all().delete()

        contact_names = request.POST.getlist('contact_name[]')
        designations = request.POST.getlist('designation[]')
        emails = request.POST.getlist('email[]')
        phone_numbers = request.POST.getlist('number[]')
        dobs = request.POST.getlist('dob[]')
        religions = request.POST.getlist('religion[]')

        for name, designation, email, phone_number, dob, religion in zip(contact_names, designations, emails, phone_numbers, dobs, religions):
            if name and designation and email and phone_number:
                # Set dob to None if it's an empty string
                if not dob:
                    dob = None
                Contact.objects.create(company=company, Contact_Name=name, designation=designation, email=email, Phone_Number=phone_number, DOB=dob, religion=religion)

        return redirect(reverse('master') + '?selection=company')
    else:
        return HttpResponse("Form Submission Error!")


def companydetails(request, company_id):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    company = Company.objects.get(pk=company_id)
    requirements = Requirement.objects.filter(company_id=company_id)
    transactions = Transaction.objects.filter(company_id=company_id).order_by('-date')
    contacts = Contact.objects.filter(company=company)

    return render(request, 'companydetails.html', {'company': company, 'requirements': requirements, 'transactions': transactions, 'contacts': contacts})

def submit_requirement(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        company_name = request.POST.get('company')
        contact_name = request.POST.get('contact-name')
        requirement_type = request.POST.get('requirement-type')
        product_name = request.POST.get('product')
        brand_name = request.POST.get('brand')
        service_name = request.POST.get('service')
        currency = request.POST.get('currency')
        price = request.POST.get('price')
        status = request.POST.get('status')
        requirement_description = request.POST.get('requirement-description')

        try:
            company = Company.objects.get(Company_Name=company_name)
        except Company.DoesNotExist:
            return HttpResponseBadRequest("Company does not exist.")
        
        try:
            contact = Contact.objects.get(Contact_Name=contact_name)
        except Contact.DoesNotExist:
            return HttpResponseBadRequest("Contact does not exist.")

        requirement = Requirement.objects.create(company=company, Contact_Name=contact, Requirement_Type=requirement_type, currency=currency, 
                                                 price=price, status=status, Requirement_Description=requirement_description)

        if requirement_type == 'Product':
            try:
                product = Product.objects.get(Product_Name=product_name)
            except Product.DoesNotExist:
                return HttpResponseBadRequest("Product does not exist.")
            
            requirement.Product_Name = product
            if brand_name:
                brand, created = Brand.objects.get_or_create(Brand_Name=brand_name)
                requirement.brand = brand

        elif requirement_type == 'Service':
            try:
                service = Service.objects.get(Service_Name=service_name)
            except Service.DoesNotExist:
                return HttpResponseBadRequest("Service does not exist.")
            
            requirement.service = service

        requirement.save()
        return redirect(reverse('master') + '?selection=requirement&success=True')
    else:
        return HttpResponseBadRequest("Invalid request method.")

@csrf_exempt
def submit_sector(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        sector_name = request.POST.get('sector')
        
        if sector_name:
            sector = Sector(Sector_Name=sector_name)
            sector.save()
            return JsonResponse({'sector_name': sector_name})
    else:
        return HttpResponse("Form Submission Error!")
    
def add_sector(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        sector_name = request.POST.get('sector')
        
        if sector_name:
            if not Sector.objects.filter(Sector_Name=sector_name).exists():
                sector = Sector(Sector_Name=sector_name)
                sector.save()
                return redirect(reverse('master') + '?selection=sector')
            else:
                return render(request, 'form.html', {
                    'error_message': 'Sector already exists!',
                    'sector_name': sector_name,
                    'formtype': 'sector'
                })
        else:
            return render(request, 'form.html', {
                'error_message': 'Sector name cannot be empty!',
                'sector_name': sector_name,
                'formtype': 'sector'
            })
    else:
        return HttpResponse("Form Submission Error!")
    
@csrf_exempt
def submit_service(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        service_name = request.POST.get('service')
        
        if service_name:
            service = Service(Service_Name=service_name)
            service.save()
            return JsonResponse({'service_name': service_name})
    else:
        return HttpResponse("Form Submission Error!")
    
def add_service(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        service_name = request.POST.get('service')
        
        if service_name:
            if not Service.objects.filter(Service_Name=service_name).exists():
                service = Service(Service_Name=service_name)
                service.save()
                return redirect(reverse('master') + '?selection=service')
            else:
                return render(request, 'form.html', {
                    'error_message': 'Service already exists!',
                    'service_name': service_name,
                    'formtype': 'service'
                })
        else:
            return render(request, 'form.html', {
                'error_message': 'Service name cannot be empty!',
                'service_name': service_name,
                'formtype': 'service'
            })
    else:
        return HttpResponse("Form Submission Error!")
    
def update_service(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        service_id = request.POST.get('serviceId')
        service_name = request.POST.get('serviceName')

        try:
            service = Service.objects.get(id=service_id)
            service.Service_Name = service_name
            service.save()
            return redirect(reverse('master') + '?selection=service')
        except Service.DoesNotExist:
            return redirect(reverse('master') + '?selection=service')

    return redirect(reverse('master') + '?selection=service')
    
@csrf_exempt
def submit_brand(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        brand_name = request.POST.get('brand')
        
        if brand_name:
            brand = Brand(Brand_Name=brand_name)
            brand.save()
            return JsonResponse({'brand_name': brand_name})
    else:
        return HttpResponse("Form Submission Error!")
    
def add_brand(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        brand_name = request.POST.get('brand')
        
        if brand_name:
            if not Brand.objects.filter(Brand_Name=brand_name).exists():
                brand = Brand(Brand_Name=brand_name)
                brand.save()
                return redirect(reverse('master') + '?selection=brand')
            else:
                return render(request, 'form.html', {
                    'error_message': 'Brand already exists!',
                    'brand_name': brand_name,
                    'formtype': 'brand'
                })
        else:
            return render(request, 'form.html', {
                'error_message': 'Brand name cannot be empty!',
                'brand_name': brand_name,
                'formtype': 'brand'
            })
    else:
        return HttpResponse("Form Submission Error!")
    
@csrf_exempt
def submit_product(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        product_name = request.POST.get('product')
        
        if product_name:
            product = Product(Product_Name=product_name)
            product.save()
            return JsonResponse({'product_name': product_name})
    else:
        return HttpResponse("Form Submission Error!")
    
def newpartner(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    return render(request, 'newpartner.html')

@csrf_exempt
def submit_newpartner(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')

    if request.method == 'POST':
        partner_name = request.POST.get('partner')
        address = request.POST.get('address')
        city = request.POST.get('city')
        state = request.POST.get('state')
        country = request.POST.get('country')
        contact_person = request.POST.get('contact')
        designation = request.POST.get('designation')
        email = request.POST.get('email')
        number = request.POST.get('number')

        if partner_name and address and city and country and contact_person and designation and email and number:
            if Partner.objects.filter(Partner_Name=partner_name).exists():
                return JsonResponse({'error': 'A partner with this name already exists.'}, status=400)

            partner = Partner(Partner_Name=partner_name, address=address, city=city, state=state, country=country, Contact_Person=contact_person, designation=designation, email=email, Phone_Number=number)
            partner.save()

            return JsonResponse({'partner_name': partner_name, 'address': address, 'city': city, 'state':state, 'country': country, 'contact_person': contact_person, 'designation': designation, 'email': email, 'number': number})
        else:
            return JsonResponse({'error': 'Missing required fields'}, status=400)
    else:
        return HttpResponse("Form Submission Error!", status=405)
    
def add_newpartner(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        partner_name = request.POST.get('partner')
        address = request.POST.get('address')
        city = request.POST.get('city')
        state = request.POST.get('state')
        country = request.POST.get('country')
        contact_person = request.POST.get('contact')
        designation = request.POST.get('designation')
        email = request.POST.get('email')
        number = request.POST.get('number')

        if partner_name:
            if not Partner.objects.filter(Partner_Name=partner_name).exists():
                partner = Partner(Partner_Name=partner_name, address=address, city=city, state=state, country=country, Contact_Person=contact_person, designation=designation, email=email, Phone_Number=number)
                partner.save()
                return redirect(reverse('master') + '?selection=partner')
            else:
                return render(request, 'newpartner.html', {
                    'error_message': 'Partner already exists!',
                    'partner_name': partner_name,
                })
        else:
            return render(request, 'newpartner.html', {
                'error_message': 'Partner name cannot be empty!',
                'partner_name': partner_name,
            })
    else:
        return redirect(reverse('master') + '?selection=partner')
    
def partnerform(request, partner_id):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    partner = Partner.objects.get(pk=partner_id)
    return render(request, 'partnerform.html', {'partner': partner})
    
def submit_partner(request):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    if request.method == 'POST':
        partner_id = request.POST.get('partner_id')
        partner = get_object_or_404(Partner, id=partner_id)

        partner_name = request.POST.get('partner')
        address = request.POST.get('address')
        city = request.POST.get('city')
        state = request.POST.get('state')
        country = request.POST.get('country')
        contact_person = request.POST.get('contact')
        designation = request.POST.get('designation')
        email = request.POST.get('email')
        number = request.POST.get('number')

        partner.Partner_Name = partner_name
        partner.address = address
        partner.city = city
        partner.state = state
        partner.country = country
        partner.Contact_Person = contact_person
        partner.designation = designation
        partner.email = email
        partner.Phone_Number = number

        partner.save()
        return redirect(reverse('master') + '?selection=partner')
    else:
        return HttpResponse("Form Submission Error!")

def partnerdetails(request, partner_id):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    partner = Partner.objects.get(pk=partner_id)
    partners = Partner.objects.all()
    return render(request, 'partnerdetails.html', { 'partner': partner, 'partners': partners})
    
def export_excel(request, company_id):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')

    company = Company.objects.get(pk=company_id)
    transactions = Transaction.objects.filter(company_id=company_id).order_by('-date')

    wb = Workbook()
    ws = wb.active

    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    bold_font = Font(bold=True)
    
    if transactions.exists():
        first_transaction = transactions.first()
        requirement = first_transaction.requirement
        requirement_type = requirement.Requirement_Type if requirement else ''
        
        if requirement_type.lower() == 'product':
            ws.append(['Requirement Type', 'Product Name', 'Company Name'])
            brand = requirement.brand.Brand_Name if requirement and requirement.brand else ''
            product_name = requirement.Product_Name.Product_Name if requirement and requirement.Product_Name else ''
            if brand and product_name:
                product_name = f"{brand} - {product_name}"
            ws.append([requirement_type, product_name, company.Company_Name])
        elif requirement_type.lower() == 'service':
            ws.append(['Requirement Type', 'Service', 'Company Name'])
            service = requirement.service.Service_Name if requirement and requirement.service else ''
            ws.append([requirement_type, service, company.Company_Name])

        ws.append([])
        ws.append(['Date', 'Action', 'Remark'])

        column_widths = {'A': 20, 'B': 50, 'C': 50}
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

        date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')
        for cell in ws['A']:
            cell.style = date_style
        
        # Apply styles to headers
        for cell in ws[1]:
            cell.alignment = center_alignment
            cell.font = bold_font
        for cell in ws[4]:
            cell.alignment = center_alignment
            cell.font = bold_font

        # Apply alignment styles to the 'Date' column
        for cell in ws['A']:
            cell.alignment = center_alignment
        for cell in ws['B']:
            cell.alignment = center_alignment
        for cell in ws['C']:
            cell.alignment = center_alignment

        # Apply alignment and styles to row values
        for row_num, transaction in enumerate(transactions, start=5):
            ws.append([transaction.date, transaction.action, transaction.remark])
            
            ws[f'A{row_num}'].alignment = center_alignment
            ws[f'B{row_num}'].alignment = left_alignment
            ws[f'C{row_num}'].alignment = left_alignment

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{company.Company_Name} - Transactions.xlsx"'
    return response

def export_pdf(request, company_id):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')
    
    company = Company.objects.get(pk=company_id)
    transactions = Transaction.objects.filter(company_id=company_id).order_by('-date')
    contacts = Contact.objects.filter(company=company)

    template_path = 'pdftemplate.html'
    context = {'company': company, 'transactions': transactions, 'contacts': contacts}
    template = get_template(template_path)
    html = template.render(context)

    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("ISO-8859-1")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{company.Company_Name} Transactions.pdf"'
        return response
    
    return HttpResponse('Error generating PDF', status=500)

def add_transaction(request, company_id, requirement_id):
    fullname = request.session.get('fullname')

    if not fullname:
        return redirect('login')

    if request.method == 'POST':
        company = get_object_or_404(Company, id=company_id)
        requirement = get_object_or_404(Requirement, id=requirement_id)
        date = request.POST['date']
        contact_id = request.POST['contact']
        contact = get_object_or_404(Contact, id=contact_id)
        action = request.POST['action']
        remark = request.POST.get('remark', '')

        Transaction.objects.create(date=date, company=company, requirement=requirement, contact=contact, action=action, remark=remark )
        return redirect('transactiondetails', company_id=company_id, requirement_id=requirement_id)